import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from tkinter import messagebox
from datetime import datetime

# Definir estilos una vez para reutilizarlos
HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN_BORDER_SIDE = Side(style='thin', color='000000')
TABLE_BORDER = Border(left=THIN_BORDER_SIDE,
                      right=THIN_BORDER_SIDE,
                      top=THIN_BORDER_SIDE,
                      bottom=THIN_BORDER_SIDE)

def format_time_for_display(time_str):
    """Convierte 'HH:MM' a formato am/pm si es posible."""
    try:
        return datetime.strptime(time_str, "%H:%M").strftime("%I:%M %p")
    except ValueError:
        return time_str # Devuelve original si hay error de formato

def get_day_of_week(date_str_iso):
    """Obtiene el nombre del día de la semana en español a partir de una fecha YYYY-MM-DD."""
    try:
        date_obj = datetime.strptime(date_str_iso, "%Y-%m-%d")
        days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        return days[date_obj.weekday()]
    except ValueError:
        return "Fecha inválida"

def export_teacher_schedule_to_excel(schedule_data, teacher_name, subject_repo, classroom_repo, file_path):
    """
    Exporta el horario de un profesor a un archivo Excel.

    Args:
        schedule_data (list): Lista de diccionarios, donde cada diccionario es una clase programada.
                              Ej: [{"date": "2024-07-20", "start_time": "08:00", "end_time": "10:00", 
                                    "subject_id": "S1", "classroom_id": "C1"}, ...]
        teacher_name (str): Nombre completo del profesor.
        subject_repo: Instancia del repositorio de materias (para obtener nombres).
        classroom_repo: Instancia del repositorio de salones (para obtener nombres).
        file_path (str): Ruta completa donde se guardará el archivo Excel.

    Returns:
        bool: True si la exportación fue exitosa, False en caso contrario.
    """
    if not schedule_data:
        messagebox.showinfo("Información", "El profesor no tiene clases programadas para exportar.", icon='info')
        return False

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Horario {teacher_name}"

    # Título del Horario
    sheet.merge_cells('A1:F1')
    title_cell = sheet['A1']
    title_cell.value = f"Horario de Clases - Profesor: {teacher_name}"
    title_cell.font = Font(name='Calibri', size=14, bold=True)
    title_cell.alignment = CENTER_ALIGNMENT
    sheet.row_dimensions[1].height = 20

    # Encabezados de la tabla
    headers = ["Día Semana", "Fecha", "Hora Inicio", "Hora Fin", "Materia", "Salón"]
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num, value=header_title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = TABLE_BORDER
        # Ajustar ancho de columnas
        if header_title == "Materia":
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 35
        elif header_title in ["Día Semana", "Fecha", "Hora Inicio", "Hora Fin", "Salón"]:
             sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18
        else:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15


    # Llenar datos del horario
    # Ordenar los datos por fecha y luego por hora de inicio
    try:
        schedule_data.sort(key=lambda x: (datetime.strptime(x.get("date", "1900-01-01"), "%Y-%m-%d"),
                                           datetime.strptime(x.get("start_time", "00:00"), "%H:%M")))
    except ValueError as e:
        print(f"Error al ordenar los datos del horario: {e}")
        # Continuar sin ordenar si hay un problema de formato, aunque es menos ideal


    current_row = 4
    for entry in schedule_data:
        date_str = entry.get("date", "N/A")
        start_time = format_time_for_display(entry.get("start_time", "N/A"))
        end_time = format_time_for_display(entry.get("end_time", "N/A"))
        
        subject_id = entry.get("subject_id")
        subject_name = subject_id
        if subject_repo and subject_id:
            subject_details = subject_repo.get_subject_by_id(subject_id)
            if subject_details:
                subject_name = subject_details[1] # Asumiendo que el nombre es el segundo elemento

        classroom_id = entry.get("classroom_id")
        classroom_name = classroom_id
        if classroom_repo and classroom_id:
            classroom_details = classroom_repo.get_classroom_by_id(classroom_id)
            if classroom_details:
                classroom_name = classroom_details[1] # Asumiendo que el nombre/código es el segundo elemento

        day_of_week = get_day_of_week(date_str)

        row_data = [
            day_of_week,
            date_str,
            start_time,
            end_time,
            subject_name,
            classroom_name
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            cell = sheet.cell(row=current_row, column=col_num, value=cell_value)
            cell.alignment = LEFT_ALIGNMENT if col_num in [1, 5, 6] else CENTER_ALIGNMENT # Alinear texto de materia y salon a la izq
            cell.border = TABLE_BORDER
        
        sheet.row_dimensions[current_row].height = 18 # Ajustar altura de fila
        current_row += 1
    
    sheet.freeze_panes = 'A4' # Congelar encabezados

    try:
        workbook.save(file_path)
        return True
    except Exception as e:
        messagebox.showerror("Error al Exportar", f"No se pudo guardar el archivo Excel.\nDetalle: {e}", icon='error')
        return False